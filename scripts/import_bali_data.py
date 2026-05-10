"""导入 02_行程产品报价 母库的两份 xlsx 到资源库.

用法:
    # dry-run:仅打印将要写入的数据
    python scripts/import_bali_data.py

    # 真写入(API 鉴权用 admin/123456,通过环境变量也可以改)
    python scripts/import_bali_data.py --apply --base http://localhost:8000

    # 只导入某一类
    python scripts/import_bali_data.py --apply --only attractions hotels

字段映射:
    景点门票 → /resources/attractions
    酒店     → /resources/hotels(主酒店 + 1 标准房型)
    车费     → /resources/vehicles(FULL DAY 价 = cost_idr_per_day)
    下午茶   → /resources/simple/tea
    瑜伽/SPA → /resources/simple/spa
    一日游   → /templates(仅名称 + 描述 + 估时,景点关联留人工)

价格清洗:
    "Rp 25,000" / "Rp25.000" → 25000
    "500k" → 500000;"500k-650k" 取低值
    "2人：450元/人" → 450 × 2300 = 1,035,000 IDR
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl
import requests

ROOT = Path(__file__).resolve().parent.parent
SRC_MAIN = Path(r"C:\Users\001\balijob\02_行程产品报价\20260410巴厘岛项目成本2026年版本.xlsx")
SRC_FRAG = Path(r"C:\Users\001\balijob\02_行程产品报价\BWS碎片化整理.xlsx")

CNY_TO_IDR = 2300  # 汇率(碎片表里"X 元"按这个换算)


# ============================================================
#  价格解析工具
# ============================================================

_RE_RP = re.compile(r"Rp\s*([0-9.,]+)", re.IGNORECASE)
_RE_K = re.compile(r"(\d+)\s*k", re.IGNORECASE)
_RE_INT = re.compile(r"(\d{1,3}(?:[.,]\d{3})+|\d+)")


def parse_idr(s: Any) -> int | None:
    """从'Rp 25,000'/'500k'/'1.220.000'等抽出数字 IDR."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return int(s)
    txt = str(s).strip()
    if not txt:
        return None

    # 含 "Rp" 前缀
    m = _RE_RP.search(txt)
    if m:
        num = m.group(1).replace(".", "").replace(",", "").replace(" ", "")
        if num.isdigit():
            return int(num)

    # 含 "k"(代表 1000)
    m = _RE_K.search(txt)
    if m:
        return int(m.group(1)) * 1000

    # 纯数字 + 逗号/句点千分位
    m = _RE_INT.search(txt)
    if m:
        num = m.group(1).replace(",", "").replace(".", "")
        if num.isdigit() and len(num) >= 4:
            return int(num)

    return None


def parse_cny_per_pax(s: Any) -> int | None:
    """从 '2人：450元/人' 抽 450."""
    if s is None:
        return None
    txt = str(s)
    m = re.search(r"(\d{2,5})\s*元", txt)
    if m:
        return int(m.group(1))
    return None


# ============================================================
#  解析器:每个 sheet 一段
# ============================================================

def parse_attractions() -> list[dict]:
    """主表 巴厘岛项目门票 → attractions."""
    wb = openpyxl.load_workbook(str(SRC_MAIN), data_only=True)
    ws = wb["巴厘岛项目门票"]
    out: list[dict] = []
    current_vendor = None
    for row in ws.iter_rows(min_row=3, values_only=True, max_col=4):
        no, vendor, desc, price = row[:4]
        if vendor and isinstance(vendor, str) and vendor.strip():
            current_vendor = vendor.strip()
        name = current_vendor
        if not name:
            continue

        # 描述纳入名称(如"漂流2小时")
        full_name = name
        if desc and isinstance(desc, str) and desc.strip() and desc.strip() != "门票":
            d = desc.strip()
            # 避免名字过长
            if len(d) < 30:
                full_name = f"{name}-{d}"

        idr = parse_idr(price)
        if not idr or idr < 1000:
            continue

        out.append({
            "destination_id": 1,
            "name_zh": full_name[:120],
            "area": None,
            "ticket_idr_adult": idr,
            "ticket_idr_child": int(idr * 0.7),  # 默认儿童 7 折
            "recommended_minutes": 60,
            "restrictions": str(desc or "") if desc else None,
        })
    return out


def parse_hotels() -> list[dict]:
    """主表 酒店 → hotels(每个 hotel 含 1 个 Standard 房型)."""
    wb = openpyxl.load_workbook(str(SRC_MAIN), data_only=True)
    ws = wb["酒店"]
    out: list[dict] = []
    for row in ws.iter_rows(min_row=4, values_only=True, max_col=4):
        star_label, hotel_name, low_price, high_addon = row[:4]
        if not (hotel_name and isinstance(hotel_name, str)):
            continue
        name = hotel_name.strip()
        if not name or "酒店名称" in name:
            continue

        # 解析星级
        star = None
        if star_label and isinstance(star_label, str):
            if "五星" in star_label or "国五" in star_label:
                star = 5
            elif "四星" in star_label:
                star = 4
            elif "三星" in star_label:
                star = 3

        # 取淡季价(最低)
        low_idr = parse_idr(low_price)
        # 旺季 = 淡季 + 加价 / 或同价
        high_idr = parse_idr(high_addon) if high_addon else None
        if high_idr and low_idr:
            high_idr = low_idr + high_idr  # "100k/间/晚"是加价
        elif not high_idr:
            high_idr = low_idr

        if not low_idr or low_idr < 100000:
            continue

        # 抽取中英文名
        # 形如 "贝斯特韦斯特卡马拉金巴兰酒店(Best Western Kamala Jimbaran)"
        m = re.match(r"^(.+?)\((.+?)\)\s*$", name)
        if m:
            zh, en = m.group(1).strip(), m.group(2).strip()
        else:
            zh, en = name, None

        out.append({
            "destination_id": 1,
            "name_zh": zh[:100],
            "name_en": en[:100] if en else None,
            "star": star,
            "area": None,
            "rooms": [{
                "room_type": "Standard / Deluxe",
                "max_occupancy": 2,
                "breakfast_included": True,
                "cost_idr_low": low_idr,
                "cost_idr_high": high_idr or low_idr,
                "supplier": "BWS 母库",
                "note": str(low_price)[:200] if low_price else None,
            }],
        })
    return out


def parse_vehicles() -> list[dict]:
    """主表 车费 → vehicles(取 FULL DAY 价).行 3-11 是 7 个标准车型."""
    wb = openpyxl.load_workbook(str(SRC_MAIN), data_only=True)
    ws = wb["车费"]
    # row 2: 表头  row 3-11: 数据  cols: 0=车型, 2=HALF, 3=FULL, 4=4D, 5=5D, 6=6D, 7=7D, 8=8D
    out: list[dict] = []
    seat_map = {
        "AVANZA": 7, "INOVA REBORN": 7, "INNOVA REBORN": 7,
        "ELF SHORT": 12, "ELF LONG": 17, "TOYOTA HIACE": 17,
        "巴士28": 28, "巴士30-35": 35, "巴士30": 30,
        "巴士40-45": 45, "巴士40": 40,
    }
    for row in ws.iter_rows(min_row=3, max_row=11, values_only=True, max_col=9):
        car_type = row[0]
        full_day = row[3] if len(row) > 3 else None
        if not (car_type and isinstance(car_type, str)):
            continue
        ct = car_type.strip()
        if not ct or "车型" in ct:
            continue
        # 决定座位
        seat = None
        for key, s in seat_map.items():
            if key in ct.upper() or key in ct:
                seat = s
                break
        if not seat:
            # 尝试从字符串提取数字
            m = re.search(r"(\d+)\s*座", ct)
            if m:
                seat = int(m.group(1))
        if not seat:
            continue
        idr = parse_idr(full_day)
        if not idr:
            continue
        out.append({
            "destination_id": 1,
            "vehicle_type": ct[:60],
            "seat_count": seat,
            "cost_idr_per_day": idr,
            "includes_fuel": True,
            "includes_driver": True,
            "terrain_note": "源:BWS 母库 2026 项目成本表",
        })
    return out


def parse_one_day_tours() -> list[dict]:
    """碎片表 巴厘岛一日游 → templates(仅 name + description + minutes)."""
    wb = openpyxl.load_workbook(str(SRC_FRAG), data_only=True)
    ws = wb["巴厘岛一日游"]
    # row 1: 表头, row 2: 提示, row 3+: 数据
    # cols: 0=NO 1=产品 2=天数 3=行程安排 4=空 5=结算价 6=有效期 7=退改 8=须知 9=注意
    out: list[dict] = []
    for row in ws.iter_rows(min_row=3, values_only=True, max_col=10):
        no, name, days, route, _, price, _, _, _, _ = row[:10]
        if not (name and isinstance(name, str)):
            continue
        nm = name.strip()
        if not nm or "以下报价" in nm:
            continue
        # 取首行作为简短名
        short_name = nm.split("\n")[0][:120]
        # 描述 = 行程安排首段
        desc = ""
        if route and isinstance(route, str):
            desc = route.strip()[:600]
        # 天数判断
        is_day_trip = True
        if days and isinstance(days, str) and "天" in days:
            try:
                d = int(re.search(r"(\d+)", days).group(1))
                if d > 1:
                    is_day_trip = False  # 多日游不算一日游模板
            except Exception:
                pass
        if not is_day_trip:
            continue
        out.append({
            "destination_id": 1,
            "name_zh": short_name,
            "description": desc,
            "total_minutes_estimate": 540,  # 默认 9 小时
            "difficulty": "easy",
            "recommended_pax_min": 2,
            "recommended_pax_max": 17,
            "attractions": [],
            "restaurants": [],
            "_source_price": str(price)[:200] if price else None,
        })
    return out


def _cny_range_avg(s: Any) -> int | None:
    """'¥210 – ¥260' / '¥1,650+/晚' → 中位整数(CNY)."""
    if s is None:
        return None
    txt = str(s).replace("，", ",").replace(",", "")
    nums = [int(m.group(1)) for m in re.finditer(r"(\d{2,5})", txt)]
    if not nums:
        return None
    if len(nums) >= 2:
        return (nums[0] + nums[1]) // 2
    return nums[0]


def parse_tea() -> list[dict]:
    """碎片表 高端下午茶 → simple/tea(列宽 6:中文名|区域|价格|亮点)."""
    wb = openpyxl.load_workbook(str(SRC_FRAG), data_only=True)
    ws = wb["高端下午茶"]
    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True, max_col=6):
        no, name_en, name_zh, location, price, highlights = row[:6] + (None,) * (6 - len(row))
        if not name_zh or not isinstance(name_zh, str):
            continue
        nm = name_zh.strip()
        if not nm or len(nm) < 3 or "📌" in nm or "Afternoon" in nm:
            continue
        cny = _cny_range_avg(price)
        if not cny or cny < 50:
            continue
        idr = cny * CNY_TO_IDR
        # 区域: "Nusa Dua 努沙杜瓦" → 取中文部分
        area = None
        if location and isinstance(location, str):
            m = re.search(r"[一-龥]{2,8}", location)
            if m:
                area = m.group(0)
        out.append({
            "destination_id": 1,
            "name_zh": nm[:120],
            "venue": str(name_en).strip()[:120] if name_en else None,
            "area": area,
            "cost_idr_per_person": idr,
            "min_pax": 2,
            "recommended_minutes": 90,
        })
    return out


def parse_spa() -> list[dict]:
    """碎片表 疗愈+瑜伽 → simple/spa(列宽 5:区域|英文名|中文|介绍|价格,区域 merged 需 carry forward)."""
    wb = openpyxl.load_workbook(str(SRC_FRAG), data_only=True)
    ws = wb["疗愈+瑜伽"]
    out: list[dict] = []
    cur_area = None
    for row in ws.iter_rows(min_row=2, values_only=True, max_col=5):
        loc, name_en, name_zh, intro, price = row[:5] + (None,) * (5 - len(row))
        # carry forward area
        if loc and isinstance(loc, str) and loc.strip():
            m = re.search(r"[一-龥]{2,8}", loc)
            if m:
                cur_area = m.group(0)
        if not name_zh or not isinstance(name_zh, str):
            continue
        nm = name_zh.strip()
        if len(nm) < 3 or "备注" in nm:
            continue
        cny = _cny_range_avg(price)
        if not cny or cny < 50:
            continue
        idr = cny * CNY_TO_IDR
        out.append({
            "destination_id": 1,
            "brand": str(name_en).strip()[:80] if name_en else nm[:80],
            "package_name": nm[:120],
            "duration_minutes": 90,
            "cost_idr_per_person": idr,
            "includes": (str(intro)[:200] + (f" / 区域:{cur_area}" if cur_area else "")) if intro else None,
        })
    return out


def parse_sub_hotels() -> list[dict]:
    """碎片表 巴厘岛酒店整理 → hotels(主推客户报价,标 supplier='BWS 客户报价表').
    多行结构:酒店名+星级+位置+房型+日期+报价+备注;merged 单元需 carry forward."""
    wb = openpyxl.load_workbook(str(SRC_FRAG), data_only=True)
    ws = wb["巴厘岛酒店整理"]
    out: list[dict] = []
    cur = None  # 当前酒店字典
    cur_star = None
    cur_area = None
    for row in ws.iter_rows(min_row=4, values_only=True, max_col=7):
        cells = list(row[:7]) + [None] * (7 - len(row))
        name_cell, star_cell, loc_cell, room_cell, date_cell, price_cell, note_cell = cells
        # 新酒店行:name_cell 非空
        if name_cell and isinstance(name_cell, str) and name_cell.strip():
            # 提交上一家
            if cur and cur["rooms"]:
                out.append(cur)
            raw_name = name_cell.strip()
            # "Episode Kuta Bali\n伊匹库塔巴厘岛酒店" or "贝斯特韦斯特...\nBest Western..."
            lines = [l.strip() for l in raw_name.split("\n") if l.strip()]
            # 中文优先选含中文字符
            zh = next((l for l in lines if re.search(r"[一-龥]", l)), lines[0])
            en = next((l for l in lines if not re.search(r"[一-龥]", l)), None)
            star = None
            if star_cell and isinstance(star_cell, str):
                if "5" in star_cell or "五" in star_cell:
                    star = 5
                elif "4" in star_cell or "四" in star_cell:
                    star = 4
                elif "3" in star_cell or "三" in star_cell:
                    star = 3
            cur_star = star
            cur_area = str(loc_cell).strip() if loc_cell else None
            cur = {
                "destination_id": 1,
                "name_zh": zh[:100],
                "name_en": en[:100] if en else None,
                "star": star,
                "area": cur_area,
                "rooms": [],
            }
        # 房型行(name 空但 room/price 有)
        if cur and room_cell and price_cell:
            try:
                cny = int(re.search(r"(\d+)", str(price_cell).replace(",", "")).group(1))
            except Exception:
                continue
            if cny < 50:
                continue
            room_lines = [l.strip() for l in str(room_cell).split("\n") if l.strip()]
            room_type = " / ".join(room_lines[:2])[:80]
            cur["rooms"].append({
                "room_type": room_type or "Standard",
                "max_occupancy": 2,
                "breakfast_included": True,
                "cost_idr_low": cny * CNY_TO_IDR,
                "cost_idr_high": cny * CNY_TO_IDR,
                "supplier": "BWS 客户报价表 2026",
                "note": (f"日期: {date_cell}; " if date_cell else "") + (str(note_cell)[:150] if note_cell else ""),
            })
    if cur and cur["rooms"]:
        out.append(cur)
    return out


def parse_premium_hotels() -> list[dict]:
    """碎片表 高端酒店 → hotels(5星级,supplier='BWS 高端报价')."""
    wb = openpyxl.load_workbook(str(SRC_FRAG), data_only=True)
    ws = wb["高端酒店"]
    out: list[dict] = []
    cur = None
    cur_star = 5
    for row in ws.iter_rows(min_row=4, values_only=True, max_col=4):
        name_cell, star_cell, room_cell, price_cell = list(row[:4]) + [None] * (4 - len(row))
        if name_cell and isinstance(name_cell, str) and name_cell.strip():
            if cur and cur["rooms"]:
                out.append(cur)
            raw = name_cell.strip()
            lines = [l.strip() for l in raw.split("\n") if l.strip()]
            zh = next((l for l in lines if re.search(r"[一-龥]", l)), lines[0])
            en_match = re.search(r"\(([^)]+)\)", raw)
            en = en_match.group(1) if en_match else None
            zh = re.sub(r"\([^)]+\)", "", zh).strip()
            cur = {
                "destination_id": 1,
                "name_zh": zh[:100],
                "name_en": en[:100] if en else None,
                "star": 5,
                "area": None,
                "rooms": [],
            }
        if cur and room_cell and price_cell:
            cny = parse_cny_per_pax(price_cell) or 0
            if not cny:
                m = re.search(r"(\d{3,5})", str(price_cell).replace(",", ""))
                if m:
                    cny = int(m.group(1))
            if cny < 100:
                continue
            cur["rooms"].append({
                "room_type": str(room_cell).strip()[:80],
                "max_occupancy": 2,
                "breakfast_included": True,
                "cost_idr_low": cny * CNY_TO_IDR,
                "cost_idr_high": cny * CNY_TO_IDR,
                "supplier": "BWS 高端报价 2026",
                "note": str(price_cell)[:120],
            })
    if cur and cur["rooms"]:
        out.append(cur)
    return out


# ============================================================
#  API 写入
# ============================================================

def _make_session() -> requests.Session:
    s = requests.Session()
    # Windows 系统代理会拦截 localhost,显式不走 proxy
    s.trust_env = False
    s.proxies = {"http": "", "https": "", "no_proxy": "localhost,127.0.0.1"}
    return s


def login(base: str) -> requests.Session:
    s = _make_session()
    r = s.post(f"{base}/api/v1/auth/login",
               json={"username": "admin", "password": "123456"}, timeout=10)
    r.raise_for_status()
    return s


def post(s: requests.Session, base: str, path: str, body: dict) -> dict:
    body = {k: v for k, v in body.items() if not k.startswith("_")}
    r = s.post(f"{base}/api/v1{path}", json=body, timeout=15)
    if r.status_code >= 400:
        return {"_error": f"{r.status_code} {r.text[:200]}"}
    return r.json()


# ============================================================
#  主流程
# ============================================================

KINDS = {
    "attractions":     ("/resources/attractions",  parse_attractions),
    "hotels":          ("/resources/hotels",       parse_hotels),
    "vehicles":        ("/resources/vehicles",     parse_vehicles),
    "tea":             ("/resources/simple/tea",   parse_tea),
    "spa":             ("/resources/simple/spa",   parse_spa),
    "templates":       ("/templates",              parse_one_day_tours),
    "sub_hotels":      ("/resources/hotels",       parse_sub_hotels),       # 客户报价主推
    "premium_hotels":  ("/resources/hotels",       parse_premium_hotels),   # 5 星高端
}


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--apply", action="store_true", help="真写入(默认 dry-run)")
    p.add_argument("--base", default="http://localhost:8000")
    p.add_argument("--only", nargs="*", choices=list(KINDS.keys()))
    p.add_argument("--limit", type=int, default=None, help="每类最多前 N 条(便于调试)")
    args = p.parse_args()

    selected = args.only or list(KINDS.keys())
    print(f"导入计划: {selected}  apply={args.apply}\n")

    sess = login(args.base) if args.apply else None

    summary: dict[str, dict] = {}
    for kind in selected:
        path, parser = KINDS[kind]
        records = parser()
        if args.limit:
            records = records[:args.limit]
        print(f"== {kind}: 解析到 {len(records)} 条 ==")
        # 打印前 3 条预览
        for r in records[:3]:
            print(f"   [SAMPLE] {json.dumps({k: v for k, v in r.items() if k != 'rooms'}, ensure_ascii=False)[:180]}")
            if r.get("rooms"):
                print(f"            rooms={len(r['rooms'])} sample: {json.dumps(r['rooms'][0], ensure_ascii=False)[:150]}")

        ok, fail = 0, 0
        if args.apply:
            for r in records:
                resp = post(sess, args.base, path, r)
                if resp.get("_error") or "id" not in resp:
                    fail += 1
                    print(f"   [FAIL] {resp.get('_error') or resp}")
                else:
                    ok += 1
            print(f"   写入 {ok} 条,失败 {fail} 条\n")
        else:
            print(f"   (dry-run, 不写入)\n")
        summary[kind] = {"parsed": len(records), "ok": ok, "fail": fail}

    print("=" * 70)
    print("汇总:")
    for k, v in summary.items():
        print(f"  {k:15s}: parsed={v['parsed']:3d}  ok={v['ok']:3d}  fail={v['fail']:3d}")


if __name__ == "__main__":
    main()

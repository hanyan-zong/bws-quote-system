"""报价单 Excel 导出 — 3 sheet: 封面汇总 / 逐日行程 / 价格明细."""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# BWS 品牌色
BRAND_BLUE = "2C5282"
BRAND_LIGHT = "EBF8FF"
BORDER_GRAY = "CBD5E0"

THIN = Side(style="thin", color=BORDER_GRAY)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _set_col_widths(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _header_row(ws, row: int, cols: list[str], fill: str = BRAND_BLUE) -> None:
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _body_cell(ws, row: int, col: int, value, *, align: str = "left", bold: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Microsoft YaHei", bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BOX


def _build_cover(ws, ctx: dict[str, Any]) -> None:
    q = ctx["quote"]
    show = ctx["show_costs"]

    ws.title = "封面汇总"
    _set_col_widths(ws, {1: 22, 2: 30, 3: 22, 4: 30})

    # 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    cell = ws.cell(row=1, column=1, value="BWS 预报价单")
    cell.font = Font(name="Microsoft YaHei", bold=True, size=20, color=BRAND_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    sub = ws.cell(
        row=2, column=1,
        value=f"报价单号: {q['quote_no']}    |    导出时间: {ctx['meta']['exported_at']}    |    导出: {ctx['meta']['exported_by']}",
    )
    sub.font = Font(name="Microsoft YaHei", size=10, color="4A5568")
    sub.alignment = Alignment(horizontal="center", vertical="center")

    # 客户/旅行社信息
    rows = [
        ("旅行社", q["agency_name"], "联系人", q["agency_contact"]),
        ("客户姓名", q["customer_name"], "目的地", q["destination_codes"]),
        ("成人 / 儿童", f"{q['pax_adult']} / {q['pax_child']}", "总人数", q["pax_total"]),
        ("出发日 — 结束日", f"{q['start_date']} → {q['end_date']}", "天数", q["total_days"]),
        ("自由活动天数", q["free_days"], "季节", q["season_label"]),
        ("客户类型", q["customer_type_label"], "首次合作", "是" if q["is_first_time_agency"] else "否"),
    ]
    if q["arrival_at"] or q["departure_at"]:
        rows.append((
            "抵达航班", f"{q['arrival_at']} {q['arrival_airport']}",
            "离开航班", f"{q['departure_at']} {q['departure_airport']}",
        ))

    r = 4
    for label1, val1, label2, val2 in rows:
        _body_cell(ws, r, 1, label1, bold=True)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=BRAND_LIGHT)
        _body_cell(ws, r, 2, val1)
        _body_cell(ws, r, 3, label2, bold=True)
        ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor=BRAND_LIGHT)
        _body_cell(ws, r, 4, val2)
        r += 1

    # 价格汇总
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    title = ws.cell(row=r, column=1, value="💰 价格汇总")
    title.font = Font(name="Microsoft YaHei", bold=True, size=14, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=BRAND_BLUE)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 28
    r += 1

    price_rows = [
        ("人均报价 ¥", f"¥ {q['price_cny_per_pax']:,.2f}"),
        ("总报价 ¥", f"¥ {q['price_cny_total']:,.2f}"),
    ]
    if show:
        price_rows.extend([
            ("CNY 总成本", f"¥ {q.get('cost_cny_total', 0):,.2f}"),
            ("IDR 总成本", f"Rp {q.get('cost_idr_total', 0):,.0f}"),
            ("人均利润 ¥", f"¥ {q.get('profit_cny_per_pax', 0):,.2f}"),
            ("人均赌额 ¥", f"¥ {q.get('gamble_cny_per_pax', 0):,.2f}"),
            ("汇率 RMB:IDR", f"1 : {q['exchange_rate']:,.2f}"),
        ])

    for label, val in price_rows:
        _body_cell(ws, r, 1, label, bold=True)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=BRAND_LIGHT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        _body_cell(ws, r, 2, val, align="right", bold=True)
        r += 1

    if not show:
        # 给客户/agent 看的免责小字
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        notice = ws.cell(
            row=r, column=1,
            value="※ 本报价为最终客户成交价, 仅供旅行社参考报给客户. 内部成本与利润数据已隐藏.",
        )
        notice.font = Font(name="Microsoft YaHei", italic=True, size=9, color="A0AEC0")
        notice.alignment = Alignment(horizontal="center", vertical="center")

    # 行程合理性 + 备注
    if ctx.get("feasibility"):
        r += 2
        _body_cell(ws, r, 1, "行程合理性校验", bold=True)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=BRAND_LIGHT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        _body_cell(ws, r, 2, ctx["feasibility"].get("label", ""))

    if q["notes"]:
        r += 1
        _body_cell(ws, r, 1, "备注", bold=True)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=BRAND_LIGHT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        _body_cell(ws, r, 2, q["notes"])
        ws.row_dimensions[r].height = 40


def _build_itinerary(ws, ctx: dict[str, Any]) -> None:
    ws.title = "逐日行程"
    headers = ["第 N 天", "日期", "类型", "酒店/房型", "用车", "导游", "餐 (午/晚)", "景点 / 体验", "备注"]
    _set_col_widths(ws, {1: 8, 2: 12, 3: 12, 4: 26, 5: 16, 6: 12, 7: 26, 8: 38, 9: 18})
    _header_row(ws, 1, headers)

    for i, d in enumerate(ctx["days"], 2):
        _body_cell(ws, i, 1, f"D{d['day_index']}", align="center", bold=True)
        _body_cell(ws, i, 2, d["date"], align="center")
        if d["is_free"]:
            type_text = f"自由 ({d['free_hours']}h)"
        elif d["free_hours"] >= 4:
            type_text = f"半自由 ({d['free_hours']}h)"
        else:
            type_text = "全程"
        _body_cell(ws, i, 3, type_text, align="center")
        _body_cell(ws, i, 4, f"{d['hotel']} / {d['room']}".strip(" /"))
        _body_cell(ws, i, 5, d["vehicle"])
        _body_cell(ws, i, 6, d["guide"])
        meals = []
        if d["breakfast_included"]:
            meals.append("含早")
        if d["lunch"]:
            meals.append(f"午: {d['lunch']}")
        if d["dinner"]:
            meals.append(f"晚: {d['dinner']}")
        if d["afternoon_tea"]:
            meals.append(f"下午茶: {d['afternoon_tea']}")
        _body_cell(ws, i, 7, "\n".join(meals))
        attrs = []
        for a in d["attractions"]:
            stay = f" ({a['stay_minutes']}min)" if a["stay_minutes"] else ""
            attrs.append(f"{a['order']}. {a['name']}{stay}")
        spa_water = []
        if d["spa"]:
            spa_water.append(f"SPA: {d['spa']}")
        if d["water_activity"]:
            spa_water.append(f"水上: {d['water_activity']}")
        full_text = "\n".join(attrs + spa_water) if (attrs or spa_water) else ""
        _body_cell(ws, i, 8, full_text)
        _body_cell(ws, i, 9, d["notes"])
        ws.row_dimensions[i].height = max(28, 14 * (1 + len(attrs) + len(spa_water) + len(meals)))


def _build_pricing(ws, ctx: dict[str, Any]) -> None:
    """价格明细 — 仅 super_admin / agency_owner 看. agent / viewer 直接给空白说明."""
    ws.title = "价格明细"
    if not ctx["show_costs"]:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        cell = ws.cell(
            row=1, column=1,
            value="🔒 当前角色无权查看 IDR 成本与利润明细\n如需查看, 请联系旅行社管理员",
        )
        cell.font = Font(name="Microsoft YaHei", size=12, color="A0AEC0", italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 60
        return

    q = ctx["quote"]
    headers = ["项目", "金额 (IDR)", "金额 (CNY)", "说明"]
    _set_col_widths(ws, {1: 22, 2: 18, 3: 16, 4: 32})
    _header_row(ws, 1, headers)

    rate = q["exchange_rate"] or 2300
    cost_idr = q.get("cost_idr_total", 0)
    cost_cny = q.get("cost_cny_total", 0)
    profit = q.get("profit_cny_per_pax", 0)
    gamble = q.get("gamble_cny_per_pax", 0)
    pax_total = q["pax_total"]

    rows = [
        ("总成本 (含全部资源)", f"Rp {cost_idr:,.0f}", f"¥ {cost_cny:,.2f}", f"汇率 1 RMB = {rate:,.2f} IDR"),
        ("人均成本", f"Rp {cost_idr / pax_total:,.0f}", f"¥ {cost_cny / pax_total:,.2f}", f"成本 ÷ 总人数 ({pax_total})"),
        ("人均利润", "—", f"¥ {profit:,.2f}", f"按客户类型 [{q['customer_type_label']}] 默认值"),
        ("人均赌额 (让利给自费)", "—", f"¥ {gamble:,.2f}", "AI 推荐, 可手动覆盖"),
        ("人均报价", "—", f"¥ {q['price_cny_per_pax']:,.2f}", "= 人均成本 + 利润 - 赌额"),
        ("总报价", "—", f"¥ {q['price_cny_total']:,.2f}", f"= 人均报价 × {pax_total} 人"),
    ]
    for i, (label, idr, cny, note) in enumerate(rows, 2):
        _body_cell(ws, i, 1, label, bold=True)
        _body_cell(ws, i, 2, idr, align="right")
        _body_cell(ws, i, 3, cny, align="right", bold=True)
        _body_cell(ws, i, 4, note)

    # 自费推荐清单
    if ctx.get("gamble") and ctx["gamble"].get("recommended_cny", 0) > 0:
        r = len(rows) + 4
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        title = ws.cell(row=r, column=1, value="🎲 配套自费推荐 (赌自费明细)")
        title.font = Font(name="Microsoft YaHei", bold=True, size=12, color="FFFFFF")
        title.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        title.alignment = Alignment(horizontal="center", vertical="center")
        r += 1
        _body_cell(ws, r, 1, "推荐让利", bold=True)
        _body_cell(ws, r, 2, f"¥ {ctx['gamble']['recommended_cny']:,.2f}", align="right")
        _body_cell(ws, r, 3, "AI 信心", bold=True)
        _body_cell(ws, r, 4, f"{(ctx['gamble'].get('ai_confidence') or 0) * 100:.0f}%", align="right")
        r += 1
        _body_cell(ws, r, 1, "判断依据", bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        _body_cell(ws, r, 2, ctx["gamble"].get("reasoning") or "")
        ws.row_dimensions[r].height = 40


def build_excel(ctx: dict[str, Any]) -> bytes:
    wb = Workbook()
    _build_cover(wb.active, ctx)
    _build_itinerary(wb.create_sheet(), ctx)
    _build_pricing(wb.create_sheet(), ctx)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

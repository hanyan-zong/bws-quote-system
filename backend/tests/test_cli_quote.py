"""bws quote 命令测试 — 重点验空表行为 + 找不到 quote 时退码."""
from __future__ import annotations


def test_quote_list_empty_db(bws):
    """空 DB 跑 list 应正常返回 0 (表头还是要打), 不能崩."""
    bws("db", "init", "--no-seed")
    r = bws("quote", "list")
    assert r.returncode == 0, f"stderr={r.stderr}"
    assert "id" in r.stdout  # 表头


def test_quote_show_missing_id_is_business_error(bws):
    bws("db", "init", "--no-seed")
    r = bws("quote", "show", "99999")
    assert r.returncode == 1, f"got {r.returncode}; stderr={r.stderr}"
    assert "BusinessError" in r.stderr
    assert "找不到" in r.stderr


def test_quote_calc_missing_id_is_business_error(bws):
    bws("db", "init", "--no-seed")
    r = bws("quote", "calc", "99999")
    assert r.returncode == 1, f"got {r.returncode}; stderr={r.stderr}"
    assert "BusinessError" in r.stderr


def test_quote_list_with_agency_filter(bws):
    """--agency 过滤参数能跑通即可 (空表)."""
    bws("db", "init", "--no-seed")
    r = bws("quote", "list", "--agency", "宜事")
    assert r.returncode == 0, f"stderr={r.stderr}"


def test_quote_list_shows_seeded_quote(bws, seed_quote):
    qid = seed_quote(quote_no="TEST-LIST-001")
    r = bws("quote", "list")
    assert r.returncode == 0, f"stderr={r.stderr}"
    assert "TEST-LIST-001" in r.stdout
    assert str(qid) in r.stdout


def test_quote_show_seeded_quote(bws, seed_quote):
    qid = seed_quote(quote_no="TEST-SHOW-001", agency_name="ShowAgency", total_days=5)
    r = bws("quote", "show", str(qid))
    assert r.returncode == 0, f"stderr={r.stderr}"
    assert "TEST-SHOW-001" in r.stdout
    assert "ShowAgency" in r.stdout
    assert "5 天" in r.stdout


def test_quote_list_agency_filter_matches(bws, seed_quote):
    seed_quote(quote_no="TEST-AG-A", agency_name="AlphaTours")
    r = bws("quote", "list", "--agency", "Alpha")
    assert r.returncode == 0, f"stderr={r.stderr}"
    assert "TEST-AG-A" in r.stdout

    r2 = bws("quote", "list", "--agency", "NoSuchAgency")
    assert r2.returncode == 0
    assert "TEST-AG-A" not in r2.stdout


def _query_quote_fields(bws, quote_id: int, *fields: str) -> dict[str, str]:
    """用 bws db query 查指定 quote 的字段值, 返回 {字段: 字符串值}."""
    cols = ", ".join(fields)
    r = bws("db", "query", f"SELECT {cols} FROM quotes WHERE id={quote_id}")
    assert r.returncode == 0, f"db query failed: {r.stderr}"
    # 输出是表格, 简单解析: 第三行是数据
    lines = [ln for ln in r.stdout.splitlines() if ln.strip()]
    # lines[0]=表头, lines[1]=分隔, lines[2]=数据
    assert len(lines) >= 3, f"unexpected db query output:\n{r.stdout}"
    values = lines[2].split()
    # 字段较多时可能有空格分歧, 但 cost/profit 等都是单 token, OK
    return dict(zip(fields, values))


def test_quote_calc_save_writes_full_fields(bws, seed_quote):
    """--save 应当走完整 recalc, 不只是 cost_* — 验证 profit/gamble/price/feasibility 都被写了."""
    qid = seed_quote(quote_no="TEST-CALC-SAVE", customer_type="family")

    # 先看初始状态: 所有 recalc 字段都应为 default 0 / 'unchecked'
    before = _query_quote_fields(
        bws, qid,
        "profit_cny_per_pax", "price_cny_per_pax", "feasibility_status",
    )
    assert before["profit_cny_per_pax"] == "0"
    assert before["feasibility_status"] == "unchecked"

    # 跑 --save
    r = bws("quote", "calc", str(qid), "--save")
    assert r.returncode == 0, f"calc --save failed: {r.stderr}"
    assert "已保存完整 recalc" in r.stdout

    # 再查: profit 必须 != 0 (family 类型默认 250), feasibility 必须不再是 unchecked
    after = _query_quote_fields(
        bws, qid,
        "profit_cny_per_pax", "price_cny_per_pax", "feasibility_status",
    )
    assert float(after["profit_cny_per_pax"]) > 0, f"profit 没被回写: {after}"
    assert after["feasibility_status"] != "unchecked", f"feasibility 没被回写: {after}"


def test_quote_calc_dry_run_does_not_persist(bws, seed_quote):
    """不带 --save 只 dry-run, 任何字段都不应落库."""
    qid = seed_quote(quote_no="TEST-CALC-DRYRUN")

    r = bws("quote", "calc", str(qid))
    assert r.returncode == 0, f"calc failed: {r.stderr}"
    assert "已保存" not in r.stdout

    after = _query_quote_fields(
        bws, qid,
        "profit_cny_per_pax", "feasibility_status",
    )
    assert after["profit_cny_per_pax"] == "0"
    assert after["feasibility_status"] == "unchecked"


# -------- export (CSV / xlsx) --------


def test_quote_export_csv_default_headers(bws, seed_quote, tmp_path):
    seed_quote(quote_no="Q-EXP-001", agency_name="AlphaTours")
    out = tmp_path / "quotes.csv"
    r = bws("quote", "export", str(out))
    assert r.returncode == 0, r.stderr
    assert out.exists()
    text = out.read_text(encoding="utf-8-sig")
    assert "报价单号" in text  # 业务友好中文表头
    assert "Q-EXP-001" in text
    assert "AlphaTours" in text


def test_quote_export_csv_has_utf8_bom(bws, seed_quote, tmp_path):
    """裸字节: 必须以 UTF-8 BOM 开头, 保证 Excel 双击不乱码."""
    seed_quote(quote_no="Q-EXP-BOM")
    out = tmp_path / "bom.csv"
    r = bws("quote", "export", str(out))
    assert r.returncode == 0, r.stderr
    assert out.read_bytes()[:3] == b"\xef\xbb\xbf"


def test_quote_export_json_summary(bws, seed_quote, tmp_path):
    import json

    seed_quote(quote_no="Q-EXP-JSON")
    out = tmp_path / "s.csv"
    r = bws("quote", "export", str(out), "--json")
    assert r.returncode == 0, r.stderr
    summary = json.loads(r.stdout)
    assert summary["exported"] == 1
    assert summary["format"] == "csv"


def test_quote_export_all_columns(bws, seed_quote, tmp_path):
    """--all 导出原始列名, 不用中文表头."""
    seed_quote(quote_no="Q-EXP-ALL")
    out = tmp_path / "all.csv"
    r = bws("quote", "export", str(out), "--all")
    assert r.returncode == 0, r.stderr
    text = out.read_text(encoding="utf-8-sig")
    assert "cost_idr_total" in text  # 原始列, 默认视图没有
    assert "报价单号" not in text


def test_quote_export_bad_extension_is_usage_error(bws, seed_quote, tmp_path):
    """未知扩展名又没 --format -> 用法错误(退码 2), 不产出文件."""
    seed_quote(quote_no="Q-EXP-BAD")
    out = tmp_path / "x.txt"
    r = bws("quote", "export", str(out))
    assert r.returncode == 2, f"stderr={r.stderr}"
    assert not out.exists()


def test_quote_export_format_override(bws, seed_quote, tmp_path):
    """--format 覆盖扩展名: .dat 也能当 csv 写."""
    seed_quote(quote_no="Q-EXP-OVR")
    out = tmp_path / "x.dat"
    r = bws("quote", "export", str(out), "--format", "csv")
    assert r.returncode == 0, r.stderr
    assert "报价单号" in out.read_text(encoding="utf-8-sig")


def test_quote_export_agency_filter_no_match(bws, seed_quote, tmp_path):
    """筛选无命中: 仍写表头, 0 数据行."""
    seed_quote(quote_no="Q-EXP-FLT", agency_name="AlphaTours")
    out = tmp_path / "none.csv"
    r = bws("quote", "export", str(out), "--agency", "NoSuchAgency")
    assert r.returncode == 0, r.stderr
    lines = [ln for ln in out.read_text(encoding="utf-8-sig").splitlines() if ln.strip()]
    assert len(lines) == 1  # 只有表头
    assert "报价单号" in lines[0]


def test_quote_export_xlsx(bws, seed_quote, tmp_path):
    """xlsx: 装了 openpyxl 则生成并可读回; 没装则退码 1 + 安装提示."""
    seed_quote(quote_no="Q-EXP-XLSX")
    out = tmp_path / "quotes.xlsx"
    r = bws("quote", "export", str(out))
    try:
        from openpyxl import load_workbook
    except ImportError:
        assert r.returncode == 1
        assert "openpyxl" in (r.stdout + r.stderr)
        return
    assert r.returncode == 0, r.stderr
    assert out.exists()
    ws = load_workbook(out).active
    headers = [c.value for c in ws[1]]
    assert "报价单号" in headers
    values = [c.value for row in ws.iter_rows(min_row=2) for c in row]
    assert "Q-EXP-XLSX" in values
